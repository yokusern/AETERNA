#!/usr/bin/env python3
"""
product_creator_agent.py
商品タイプ別ファイル生成エージェント
notion_template / prompt_pack / script_pack / guide /
spreadsheet_pack / html_tool / checklist_pack / business_template
"""
import json, zipfile
from pathlib import Path
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PRODUCTS_DIR = Path(__file__).parent.parent / "products"

# ─── Notion Template ────────────────────────────────────────────────────────

def create_notion_template(spec: dict, product_dir: Path):
    name  = spec["name"]
    topic = spec.get("subcategory", "Productivity")
    price = spec.get("price_usd", 19)

    template_md = f"""# {name} — Notion Template Pack

> **Setup:** 15 minutes  |  **Works with:** Notion Free & Paid plans

---

## What's Included

A complete {topic} system you can duplicate into your Notion workspace.

---

## 1. Dashboard (Home)

Central hub showing key metrics, quick links, and today's agenda.

Properties:
- Date (Date)
- Status (Select: Today / This Week / Someday)
- Priority (Select: High / Medium / Low)

---

## 2. Projects Database

Track every project from idea to done.

| Property | Type | Options |
|---|---|---|
| Project Name | Title | — |
| Status | Select | Planning / Active / On Hold / Done |
| Priority | Select | P1 / P2 / P3 |
| Due Date | Date | — |
| Progress | Number | 0–100% |
| Tags | Multi-select | — |

Views: Board (by Status), Timeline (by Due Date), Table, Gallery

---

## 3. Tasks Database

| Property | Type | Options |
|---|---|---|
| Task Name | Title | — |
| Project | Relation → Projects | — |
| Status | Select | Todo / In Progress / Blocked / Done |
| Due Date | Date | — |
| Estimated Time | Number | hours |
| Priority | Select | High / Medium / Low |

Views: Today, This Week, By Project, Kanban

---

## 4. Goals Tracker

Annual → Quarterly → Monthly goal cascade.

| Property | Type | Description |
|---|---|---|
| Goal | Title | The goal statement |
| Type | Select | Annual / Quarterly / Monthly |
| Target Date | Date | — |
| Key Result | Text | Measurable success criteria |
| Progress | Number | 0–100% |
| Status | Select | On Track / At Risk / Achieved |

---

## 5. Resources / Knowledge Base

| Property | Type | Description |
|---|---|---|
| Title | Title | — |
| Type | Select | Article / Book / Video / Tool |
| URL | URL | — |
| Tags | Multi-select | — |
| Status | Select | Saved / Reading / Done |
| Summary | Text | Key takeaways |

---

## 6. Weekly Review Template

- What did I accomplish this week?
- What didn't get done? (why?)
- Top 3 priorities for next week
- One thing I learned
- Gratitude note

---

## Setup Guide (Step by Step)

1. Open Notion → New Page → name it "{name}"
2. Create each database (type `/database`)
3. Add properties from the tables above
4. Import the CSV files: `database → ⋯ → Import → CSV`
5. Create relations: Tasks → Projects (Relation property)
6. Link databases to Dashboard using Linked Views

---

## Pro Tips

- Use **Notion Web Clipper** to save articles to Resources
- Set **recurring reminders** for Weekly Review (every Sunday)
- Share individual pages with teammates via Share → Copy link
- Use **Gallery view** with cover images for a visual project board

---

*© AETERNA Holdings — Free to modify for personal/commercial use*
"""

    projects_csv = "Project Name,Status,Priority,Due Date,Progress,Tags\nLaunch product,Active,P1,2026-06-30,45,Marketing\nQ2 planning,Done,P1,2026-04-01,100,Strategy\nWebsite redesign,Planning,P2,2026-07-15,10,Design\n"
    tasks_csv    = "Task Name,Status,Priority,Due Date,Estimated Time\nDraft proposal,Todo,High,2026-05-20,2\nReview analytics,In Progress,High,2026-05-17,1\nTeam sync,Done,Medium,2026-05-15,0.5\n"
    goals_csv    = "Goal,Type,Target Date,Key Result,Progress,Status\nGrow revenue 50%,Annual,2026-12-31,MRR > $10k,30,On Track\nLaunch 3 products,Quarterly,2026-06-30,3 live products,33,On Track\n"
    setup_md     = f"# Quick Start — {name}\n\n1. Open Notion\n2. Create page '{name}'\n3. Build each database per the spec\n4. Import CSV files\n5. Customize and enjoy!\n"

    (product_dir / "template-spec.md").write_text(template_md, encoding="utf-8")
    (product_dir / "projects.csv").write_text(projects_csv, encoding="utf-8")
    (product_dir / "tasks.csv").write_text(tasks_csv, encoding="utf-8")
    (product_dir / "goals.csv").write_text(goals_csv, encoding="utf-8")
    (product_dir / "setup-guide.md").write_text(setup_md, encoding="utf-8")

    zip_path = product_dir / f"{spec['product_id']}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in product_dir.glob("*.md"):  zf.write(f, f.name)
        for f in product_dir.glob("*.csv"): zf.write(f, f.name)
    return zip_path


# ─── Prompt Pack ─────────────────────────────────────────────────────────────

def create_prompt_pack(spec: dict, product_dir: Path):
    name  = spec["name"]
    topic = spec.get("subcategory", "AI Productivity")

    prompts_md = f"""# {name}
## 100 Battle-Tested Prompts for {topic}

> Works with ChatGPT-4o, Claude 4, Gemini 2, and any modern LLM.
> Copy → Fill [BRACKETS] → Get results.

---

## SECTION 1: Content Creation (1–20)

**1. Blog Post Writer**
```
Write a detailed, SEO-optimized blog post about [TOPIC].
Target keyword: [KEYWORD] | Word count: [1500/2000] words | Tone: [professional/casual]
Include: intro hook, 5 H2 sections with actionable tips, conclusion with CTA.
Audience: [describe your reader]
```

**2. Social Media Caption Pack**
```
Create 10 Instagram captions for [PRODUCT/SERVICE].
Each: 150–200 chars, 1 engagement question, 5 relevant hashtags.
Vary tone: inspiring, educational, behind-the-scenes, promotional, storytelling.
```

**3. Email Subject Lines**
```
Generate 20 subject lines for [newsletter/sale/launch] campaign.
Mix: curiosity, urgency, benefit, question styles. Under 50 chars each.
Context: [audience] [offer]
```

**4. YouTube Script Outline**
```
Script outline for: "[VIDEO TITLE]"
- Hook (30 sec) - Problem setup (60 sec) - [N] main points with examples
- Transitions between sections - CTA and outro
Target: [8/10/15] minutes
```

**5. Viral X/Twitter Thread**
```
Write a viral thread about [TOPIC].
Hook tweet: bold claim. 8–10 follow-up tweets. Max 280 chars each.
End with strong CTA. Style: [educational/storytelling/how-to]
```

**6. LinkedIn Post**
```
Write a LinkedIn post about [TOPIC/EXPERIENCE].
Hook: 2–3 line personal story. Body: key insight in short paragraphs.
Takeaway: 3 lessons. CTA: question for comments. 300–500 words.
```

**7. Sales Page Copy**
```
Long-form sales page for [PRODUCT].
Sections: Headline → Pain points (5 bullets) → Story → Solution →
Feature+Benefit list (5) → Testimonials placeholder → FAQ (5) → Guarantee → 3x CTA
```

**8. Cold Email Sequence (5 emails)**
```
5-email cold outreach for [OFFER].
Day 1: Value intro (no pitch) | Day 3: Case study | Day 5: Soft pitch |
Day 8: Objection handling | Day 11: Final + scarcity
Each: subject line + 150-word body
```

**9. Product Description**
```
Compelling product description for [PRODUCT NAME].
Features: [list 3–5] | Customer: [describe] | Benefit: [transformation]
Tone: [persuasive/informative] | Length: 200 words
Include: headline, 3 bullets, closing sentence
```

**10. Newsletter Issue**
```
Weekly newsletter for [BRAND/NICHE].
1. Opening note (personal, 2–3 sentences)
2. Main article: [TOPIC] (400 words, actionable)
3. 3 quick tips
4. Resource of the week
5. Closing CTA
```

**11–20: Additional Content Prompts**
11. Press release writer
12. Podcast episode outline
13. Ad copy (FB/Google/TikTok) — 3 variations
14. Case study (challenge → solution → results)
15. Testimonial request email
16. About page writer
17. FAQ generator (15 questions)
18. Comparison article ([A] vs [B])
19. Video description + 20 SEO tags
20. Bio writer (50/100/200 word versions)

---

## SECTION 2: Business & Strategy (21–40)

**21. SWOT Analysis**
```
Detailed SWOT for [COMPANY/PRODUCT].
5 specific points per quadrant + strategic implication.
Format: structured table + 200-word strategic summary.
```

**22. Pricing Strategy**
```
Help me price [PRODUCT/SERVICE].
Cost: $[X] | Competitors: $[range] | Customer budget: $[range] | Value: [outcome]
Recommend 3 tiers (names, prices, inclusions). Justify with psychological pricing.
```

**23. Customer Persona**
```
Detailed persona for [PRODUCT].
Demographics, psychographics, pain points (top 3), goals (1 year),
buying behavior (research, trust signals, price sensitivity), authentic quote.
```

**24. OKR Framework**
```
OKRs for [TEAM/QUARTER].
3 Objectives (inspiring, qualitative).
3 Key Results each (measurable, time-bound).
Include: confidence 1–10, owner, tracking cadence.
```

**25. 30-Day Launch Strategy**
```
Product launch plan for [PRODUCT].
Week 1: Pre-launch | Week 2: Launch | Week 3: Post-launch | Week 4: Optimize
Daily actions, platforms, metrics to track.
```

**26–40: Additional Business Prompts**
26. Market research brief
27. Competitive analysis (5 competitors)
28. Revenue model options (3 models)
29. Pitch deck outline (10 slides)
30. Annual report executive summary
31. Board meeting agenda
32. Partnership proposal
33. Investor update email
34. Unit economics calculator
35. Job description writer
36. Performance review template
37. Budget justification memo
38. Hiring scorecard
39. Team meeting retrospective
40. Exit strategy options

---

## SECTION 3: Productivity (41–60)

**41. Weekly Planner**
```
Act as my productivity coach. Plan my week.
Priorities: [LIST] | Hours available: [X/day Mon-Fri]
Recurring commitments: [LIST]
Create: daily blocks, time estimates, buffer time, EOD checklist
```

**42. Decision Framework**
```
Help me decide: [DECISION]. Options: [A vs B].
My values: [LIST] | Constraints: [budget/time/risk]
Apply: pros/cons matrix, 10/10/10 rule, second-order effects.
Recommend clearly with reasoning.
```

**43. 30-Day Learning Plan**
```
Learning plan for [SKILL]. Level: [beginner/intermediate].
Time: [X hours/week] | Style: [reading/watching/doing]
Week-by-week curriculum, best resources, milestones, final project.
```

**44–60: More Productivity Prompts**
44. Morning routine builder
45. 30 journaling prompts by theme
46. Habit tracker system design
47. Focus session planner
48. Email inbox zero system
49. Weekly review template
50. Annual life audit
51. Skill gap analyzer
52. Energy management plan
53. Accountability system design
54. Reading list curator
55. Network mapping exercise
56. Career roadmap (3 years)
57. Personal finance audit
58. Meeting agenda optimizer
59. Work-life balance audit
60. "Second brain" setup guide

---

## SECTION 4: AI & Tech (61–80)

**61. Code Review**
```
Review this [LANGUAGE] code for: bugs, performance, security, readability.
[PASTE CODE]
Output: specific line feedback + refactored version.
```

**62. Bug Finder**
```
Bug in my [LANGUAGE] code.
Expected: [X] | Actual: [Y] | Error: [paste]
Code: [paste]
Diagnose root cause + provide fix with explanation.
```

**63. System Design**
```
Design system for [APP].
Scale: [users/requests] | Features: [list] | Budget: [constraint]
Output: architecture, tech stack, DB schema, API design, scaling plan.
```

**64–80: Additional Tech Prompts**
64. README generator
65. API documentation writer
66. User story writer (Agile)
67. Database schema designer
68. Regular expression builder
69. Security audit checklist
70. Tech stack recommender
71. CI/CD pipeline designer
72. Performance optimization guide
73. Error message explainer
74. Algorithm explainer (ELI5)
75. Unit test writer
76. DevOps runbook creator
77. Data analysis assistant
78. AI model selection guide
79. Microservices architecture
80. Code commenter

---

## SECTION 5: Creative (81–100)

**81–100: Creative Prompts**
81. Story plot generator (any genre)
82. Character development sheet
83. Dialogue writer
84. World-building framework
85. Poetry (5 styles)
86. Song lyrics writer
87. Movie pitch developer
88. Brand name generator (20 options)
89. Tagline creator (10 variations)
90. Logo design brief
91. Color palette rationale
92. UI microcopy writer
93. Onboarding flow copy
94. Error message humanizer
95. Legal disclaimer simplifier
96. Translation localization guide
97. Accessibility copy checker
98. Voice & tone guide
99. Style guide creator
100. Brand story writer

---

## How to Get Best Results

1. **Fill ALL [BRACKETS]** — vague input = vague output
2. **Add examples** — "Write like [specific author/brand]"
3. **Specify format** — "Output as JSON / bullet list / table"
4. **Iterate** — treat first output as a draft
5. **Chain prompts** — Prompt 23 output → into Prompt 7 → into Prompt 8

**Chain Example:**
Step 1: Prompt 23 → Customer Persona
Step 2: Paste persona → Prompt 7 → Sales Page
Step 3: Paste page → Prompt 8 → Cold Emails
= Full funnel in 20 minutes

---
*© AETERNA Holdings — Use prompts for personal and commercial projects.*
"""

    (product_dir / "100-prompts.md").write_text(prompts_md, encoding="utf-8")

    zip_path = product_dir / f"{spec['product_id']}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in product_dir.glob("*.md"):  zf.write(f, f.name)
        for f in product_dir.glob("*.txt"): zf.write(f, f.name)
    return zip_path


# ─── Script Pack ─────────────────────────────────────────────────────────────

def create_script_pack(spec: dict, product_dir: Path):
    name = spec["name"]

    readme = f"""# {name} — Python Automation Scripts

10 ready-to-use scripts. pip install → run.

| Script | What it does |
|---|---|
| 01_bulk_rename.py | Rename files with regex patterns |
| 02_csv_cleaner.py | Deduplicate, fix encoding, standardize CSVs |
| 03_web_scraper.py | Scrape any site to CSV |
| 04_folder_organizer.py | Auto-sort files by type + date |
| 05_image_resizer.py | Batch resize/compress images |
| 06_json_to_excel.py | Convert JSON API data to Excel |
| 07_text_extractor.py | Extract text from PDFs |
| 08_api_poller.py | Monitor any endpoint, alert on changes |
| 09_email_sender.py | Send bulk emails from CSV list |
| 10_report_generator.py | Auto-generate reports from CSV data |

```bash
pip install -r requirements.txt
python 01_bulk_rename.py --help
```
"""

    requirements = "requests>=2.31.0\npandas>=2.0.0\nopenpyxl>=3.1.0\nPyPDF2>=3.0.0\nPillow>=10.0.0\nbeautifulsoup4>=4.12.0\n"

    s01 = '''#!/usr/bin/env python3
"""01_bulk_rename.py - Batch file renamer with regex"""
import argparse, re
from pathlib import Path

def rename(folder, pattern, replacement, dry_run=True):
    count = 0
    for f in Path(folder).iterdir():
        if not f.is_file(): continue
        new = re.sub(pattern, replacement, f.name)
        if new != f.name:
            if dry_run: print(f"[DRY] {f.name} -> {new}")
            else: f.rename(f.parent / new); print(f"Renamed: {f.name} -> {new}")
            count += 1
    print(f"{'Would rename' if dry_run else 'Renamed'} {count} files.")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("folder"); p.add_argument("--pattern", default=" ")
    p.add_argument("--replace", default="_"); p.add_argument("--go", action="store_true")
    a = p.parse_args(); rename(a.folder, a.pattern, a.replace, not a.go)
'''

    s03 = '''#!/usr/bin/env python3
"""03_web_scraper.py - Scrape any website to CSV"""
import argparse, csv, requests
from bs4 import BeautifulSoup

def scrape(url, selector, output, limit=200):
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
    r.raise_for_status()
    items = BeautifulSoup(r.text, "html.parser").select(selector)[:limit]
    rows = [{"text": el.get_text(strip=True), "href": el.get("href",""), "src": el.get("src","")} for el in items]
    with open(output, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["text","href","src"]); w.writeheader(); w.writerows(rows)
    print(f"Saved {len(rows)} items -> {output}")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("url"); p.add_argument("--selector", default="a")
    p.add_argument("--output", default="scraped.csv"); p.add_argument("--limit", type=int, default=200)
    a = p.parse_args(); scrape(a.url, a.selector, a.output, a.limit)
'''

    s04 = '''#!/usr/bin/env python3
"""04_folder_organizer.py - Auto-organize files by type and date"""
import shutil, argparse
from pathlib import Path
from datetime import datetime

TYPES = {
    "Images":    {".jpg",".jpeg",".png",".gif",".webp",".heic"},
    "Docs":      {".pdf",".doc",".docx",".txt",".md",".xlsx",".csv"},
    "Videos":    {".mp4",".mov",".avi",".mkv"},
    "Audio":     {".mp3",".wav",".flac",".aac"},
    "Code":      {".py",".js",".ts",".html",".css",".json"},
    "Archives":  {".zip",".tar",".gz",".rar"},
}

def organize(folder, by_date=False, dry_run=True):
    count = 0
    for f in Path(folder).iterdir():
        if not f.is_file() or f.name.startswith("."): continue
        cat = next((c for c,exts in TYPES.items() if f.suffix.lower() in exts), "Other")
        dest = Path(folder)/cat/(datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m") if by_date else "")
        if dry_run: print(f"[DRY] {f.name} -> {cat}/")
        else: dest.mkdir(parents=True, exist_ok=True); shutil.move(str(f), dest/f.name); print(f"Moved: {f.name}")
        count += 1
    print(f"{'Would move' if dry_run else 'Moved'} {count} files.")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("folder"); p.add_argument("--by-date", action="store_true")
    p.add_argument("--go", action="store_true")
    a = p.parse_args(); organize(a.folder, a.by_date, not a.go)
'''

    (product_dir / "README.md").write_text(readme, encoding="utf-8")
    (product_dir / "requirements.txt").write_text(requirements, encoding="utf-8")
    (product_dir / "01_bulk_rename.py").write_text(s01, encoding="utf-8")
    (product_dir / "03_web_scraper.py").write_text(s03, encoding="utf-8")
    (product_dir / "04_folder_organizer.py").write_text(s04, encoding="utf-8")

    zip_path = product_dir / f"{spec['product_id']}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in product_dir.iterdir():
            if f.suffix in (".py",".md",".txt"): zf.write(f, f.name)
    return zip_path


# ─── Guide / Ebook ───────────────────────────────────────────────────────────

def create_guide(spec: dict, product_dir: Path):
    name  = spec["name"]
    topic = spec.get("subcategory", "Productivity")

    guide = f"""# {name}
### The Complete Practical Guide to {topic}

---

## Introduction

This guide gives you everything you need about {topic} — from first principles
to advanced strategies used by practitioners today.

**By the end you will:**
- Understand the core concepts deeply
- Apply practical techniques immediately
- Avoid the 5 most common mistakes
- Have a system you can maintain long-term

---

## Part 1: Foundations

### What Is {topic} and Why It Matters Now

{topic} is one of the highest-leverage areas you can invest time in.
The reason: small, consistent inputs compound into large outputs over time.
Most people underestimate this. The ones who don't are the ones who succeed.

### The 3 Core Principles

**Principle 1 — Outcome before method**
Define what success looks like for you, then choose the method that fits.
Not the other way around.

**Principle 2 — 80/20 your effort**
20% of the techniques in {topic} produce 80% of the results.
This guide focuses exclusively on that 20%.

**Principle 3 — Systems over willpower**
A system that works on your worst day beats a perfect plan you only follow
on good days. We build systems here.

---

## Part 2: The 5 Core Techniques

### Technique 1: The Foundation Setup
**What:** Establishing the baseline environment for {topic} to work.
**Why it works:** Without the right context, even the best techniques fail.
**How to apply:**
1. Clear your existing setup — remove what's not serving you
2. Define your one primary metric (the number that tells you it's working)
3. Set up a simple daily tracking method (notebook, app, or spreadsheet)
**Time investment:** 2 hours once, then 5 minutes daily
**Expected result:** Clarity on where you stand within 1 week

### Technique 2: The Daily Practice
**What:** The core habit that compounds over time.
**Why it works:** Frequency beats intensity. Small daily input > occasional bursts.
**How to apply:**
1. Schedule a fixed 20-minute block daily (same time is best)
2. Use the "minimum viable session" rule: even 5 minutes counts
3. Track your streak — don't break the chain
**Time investment:** 20 minutes/day
**Expected result:** Noticeable progress in 21 days, significant in 90

### Technique 3: The Weekly Review
**What:** A 30-minute weekly audit that keeps you on course.
**Why it works:** Without reflection, you optimize the wrong things.
**How to apply — every Sunday:**
- Review last week: what worked, what didn't
- Plan next week: 3 priorities maximum
- Adjust your system based on data
**Time investment:** 30 minutes/week
**Expected result:** 40% fewer wasted hours

### Technique 4: The Deep Dive
**What:** Monthly 2-hour sessions for strategic improvement.
**Why it works:** Daily habits handle the routine. Deep dives handle the growth.
**How to apply:**
- Block 2 hours on the first Sunday of each month
- Review your metric trend over the past 30 days
- Identify the one constraint limiting your progress most
- Design an experiment to address it
**Time investment:** 2 hours/month
**Expected result:** Continuous improvement without plateaus

### Technique 5: The Community Loop
**What:** Regular interaction with others working on the same thing.
**Why it works:** Shared accountability multiplies commitment by 3–5x.
**How to apply:**
- Find 1 accountability partner or join an online community
- Share your weekly result publicly (text, post, or DM)
- Give feedback to others — teaching accelerates your own learning
**Time investment:** 30 minutes/week
**Expected result:** Higher consistency, faster problem-solving

---

## Part 3: Advanced Strategies

### Strategy 1: Technique Stacking
Once techniques 1–3 are automatic (usually weeks 4–6), combine them:

**Morning block (25 min):** Technique 2 → Technique 1 check-in
**End of workday (10 min):** Technique 2 wrap-up
**Sunday (60 min):** Technique 3 full review + Technique 4 if month-end

This creates a self-reinforcing loop where each session feeds the next.

### Strategy 2: The Constraint Method
At any point, one factor limits your progress more than anything else.
Find it → fix it → repeat.

Framework:
1. Look at your metric trend — where does it plateau?
2. List 3 possible causes
3. Test one change for 2 weeks
4. Measure the result
5. Keep what works, discard what doesn't

### Strategy 3: Deliberate Discomfort
Progress stops when things get comfortable. Build in a monthly challenge:
- Something that's 20% harder than your current level
- A format you haven't tried before
- Teaching someone else what you know

---

## Part 4: Common Mistakes

### Mistake 1: Starting too big
**The trap:** Building a complex system before validating the basics.
**The fix:** Use Technique 1 only for your first 2 weeks. Add complexity only after the foundation holds.

### Mistake 2: Measuring activity instead of outcome
**The trap:** "I did the thing for 30 minutes" without checking if it moved the metric.
**The fix:** Track your one primary metric weekly. Activity is an input, not a result.

### Mistake 3: All-or-nothing thinking
**The trap:** Missing one day and calling the week a failure.
**The fix:** The "never miss twice" rule. One miss is random. Two is a pattern. The goal is never-miss-twice, not perfection.

### Mistake 4: Copying someone else's system exactly
**The trap:** A system designed for someone else's life, schedule, and goals rarely fits yours.
**The fix:** Start with someone else's framework. Customize within 30 days.

### Mistake 5: Quitting before the compound curve
**The trap:** Stopping at week 3–4, right before the compounding begins.
**The fix:** Commit to 90 days minimum before evaluating. The first 30 days are investment. Days 60–90 are when you start to see returns.

---

## Part 5: Your 30-Day Action Plan

### Week 1: Setup (Days 1–7)
- [ ] Read this guide fully
- [ ] Implement Technique 1 (setup)
- [ ] Define your primary metric
- [ ] Start daily tracking
- [ ] Write 1 sentence each evening: "Today I [action] and noticed [result]"

### Week 2: Consistency (Days 8–14)
- [ ] Technique 2: complete every day
- [ ] First weekly review (Technique 3)
- [ ] Find 1 accountability partner
- [ ] Block weekly review time in calendar

### Week 3: Deepen (Days 15–21)
- [ ] Add Technique 5 (community)
- [ ] First experiment based on weekly review
- [ ] Share progress publicly once
- [ ] Notice what's getting easier — that's the compound curve starting

### Week 4: Systematize (Days 22–30)
- [ ] First monthly deep dive (Technique 4)
- [ ] Finalize your personal version of the system
- [ ] Set 90-day target for your metric
- [ ] Celebrate: you've built a real practice

---

## Tools

| Need | Free Option | Paid Option |
|---|---|---|
| Tracking | Notion free | Notion Plus |
| Timer | Forest | Focus@Will |
| Community | Reddit / Discord | Mastermind group |
| Notes | Obsidian | Roam Research |

---

## Conclusion

The system in this guide is designed to work — if you work it.
The techniques are simple. The consistency is the hard part.

Everyone who succeeds at {topic} has one thing in common:
they kept going when it felt like nothing was happening.
That phase is actually when the compound curve is building.

**Start today. 20 minutes. That's all.**

---
*© AETERNA Holdings — Licensed for personal and commercial use.*
"""
    (product_dir / "content.md").write_text(guide, encoding="utf-8")
    return product_dir / "content.md"


# ─── Spreadsheet Pack (Excel .xlsx) ─────────────────────────────────────────

def _xl_header_style(ws, row, cols, fill_hex="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")

def _xl_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def create_spreadsheet_pack(spec: dict, product_dir: Path):
    if not HAS_OPENPYXL:
        return create_guide(spec, product_dir)

    name  = spec["name"]
    pid   = spec["product_id"]
    topic = spec.get("subcategory", "Finance")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Monthly Budget Tracker ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Budget Tracker"
    ws1.sheet_view.showGridLines = True

    _xl_header_style(ws1, 1, 5)
    headers = ["Category", "Budgeted ($)", "Actual ($)", "Difference ($)", "Status"]
    for i, h in enumerate(headers, 1):
        ws1.cell(1, i).value = h

    rows = [
        ("Housing",      1500, 0), ("Food",          600, 0),
        ("Transport",     300, 0), ("Utilities",      200, 0),
        ("Entertainment", 150, 0), ("Subscriptions",  100, 0),
        ("Health",        200, 0), ("Savings",        500, 0),
        ("Emergency Fund",200, 0), ("Misc",           150, 0),
    ]
    for r, (cat, bud, act) in enumerate(rows, 2):
        ws1.cell(r, 1).value = cat
        ws1.cell(r, 2).value = bud
        ws1.cell(r, 3).value = act
        ws1.cell(r, 4).value = f"=B{r}-C{r}"
        ws1.cell(r, 5).value = f'=IF(D{r}>=0,"✓ Under","⚠ Over")'

    total_row = len(rows) + 2
    ws1.cell(total_row, 1).value = "TOTAL"
    ws1.cell(total_row, 1).font = Font(bold=True)
    for col in range(2, 5):
        ws1.cell(total_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})"
        ws1.cell(total_row, col).font = Font(bold=True)

    _xl_col_widths(ws1, [20, 14, 14, 16, 12])

    # ── Sheet 2: Habit Tracker ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Habit Tracker")
    days = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun","Week Total","Month %"]
    ws2.cell(1, 1).value = "Habit"
    ws2.cell(1, 1).font = Font(bold=True)
    for i, d in enumerate(days, 2):
        ws2.cell(1, i).value = d
        ws2.cell(1, i).font = Font(bold=True)

    habits = ["Morning workout","Read 30 min","Meditate","No junk food",
              "8 glasses water","Journal","Study/Learn","No phone before bed"]
    for r, habit in enumerate(habits, 2):
        ws2.cell(r, 1).value = habit
        for col in range(2, 9):
            ws2.cell(r, col).value = ""
        ws2.cell(r, 9).value = f"=COUNTIF(B{r}:H{r},\"✓\")"
        ws2.cell(r, 10).value = f"=G{r}/7*100"

    _xl_col_widths(ws2, [22]+[8]*7+[12,10])

    # ── Sheet 3: Income & Expense Log ───────────────────────────────────────
    ws3 = wb.create_sheet("Transaction Log")
    _xl_header_style(ws3, 1, 5, "145A32")
    for i, h in enumerate(["Date","Description","Category","Amount ($)","Type"], 1):
        ws3.cell(1, i).value = h
    sample = [
        ("2026-05-01","Salary","Income",3500,"Income"),
        ("2026-05-03","Rent","Housing",-1500,"Expense"),
        ("2026-05-05","Groceries","Food",-120,"Expense"),
        ("2026-05-07","Freelance payment","Income",800,"Income"),
    ]
    for r, row in enumerate(sample, 2):
        for c, val in enumerate(row, 1):
            ws3.cell(r, c).value = val
    _xl_col_widths(ws3, [14, 28, 16, 14, 10])

    # ── Sheet 4: Savings Goals ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Savings Goals")
    _xl_header_style(ws4, 1, 5, "7B3F00")
    for i, h in enumerate(["Goal","Target ($)","Saved ($)","Remaining","Progress %"], 1):
        ws4.cell(1, i).value = h
    goals = [("Emergency Fund",5000,1200),("Vacation",3000,500),
              ("New laptop",1500,800),("Investment",10000,2000)]
    for r, (g, t, s) in enumerate(goals, 2):
        ws4.cell(r,1).value = g
        ws4.cell(r,2).value = t
        ws4.cell(r,3).value = s
        ws4.cell(r,4).value = f"=B{r}-C{r}"
        ws4.cell(r,5).value = f"=ROUND(C{r}/B{r}*100,1)"
    _xl_col_widths(ws4, [22,12,12,12,12])

    xlsx_path = product_dir / f"{pid}.xlsx"
    wb.save(xlsx_path)

    readme = f"""# {name} — Spreadsheet Pack

## What's Included

| Sheet | Purpose |
|---|---|
| Budget Tracker | Monthly budget vs actual with status indicators |
| Habit Tracker | 7-day habit grid with weekly totals |
| Transaction Log | Full income & expense log |
| Savings Goals | Goal tracking with progress % |

## How to Use

1. Open `{pid}.xlsx` in Excel, Google Sheets, or LibreOffice
2. **Budget Tracker**: Fill in your actual spending in column C each month
3. **Habit Tracker**: Type ✓ in daily cells when you complete a habit
4. **Transaction Log**: Add each transaction as you go
5. **Savings Goals**: Update "Saved ($)" regularly to track progress

## Compatible With
- Microsoft Excel 2016+
- Google Sheets (upload directly)
- LibreOffice Calc

*© AETERNA Holdings*
"""
    (product_dir / "README.md").write_text(readme, encoding="utf-8")

    zip_path = product_dir / f"{pid}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(xlsx_path, xlsx_path.name)
        zf.write(product_dir / "README.md", "README.md")
    return zip_path


# ─── HTML Tool ───────────────────────────────────────────────────────────────

def create_html_tool(spec: dict, product_dir: Path):
    name  = spec["name"]
    pid   = spec["product_id"]
    topic = spec.get("subcategory", "Productivity")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{name}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0f0f1a;color:#e2e8f0;min-height:100vh;padding:20px}}
  .header{{text-align:center;padding:40px 20px 20px;}}
  h1{{font-size:2rem;font-weight:700;background:linear-gradient(135deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:8px}}
  .subtitle{{color:#94a3b8;font-size:1rem}}
  .card{{background:#1e1e2e;border:1px solid #2d2d3f;border-radius:16px;padding:24px;margin:16px auto;max-width:700px}}
  .card h2{{font-size:1.1rem;font-weight:600;color:#a78bfa;margin-bottom:16px;display:flex;align-items:center;gap:8px}}
  label{{display:block;font-size:.85rem;color:#94a3b8;margin-bottom:4px;margin-top:12px}}
  input,select{{width:100%;padding:10px 14px;background:#2d2d3f;border:1px solid #3d3d5f;border-radius:8px;color:#e2e8f0;font-size:1rem;outline:none}}
  input:focus,select:focus{{border-color:#7c3aed}}
  .btn{{width:100%;padding:14px;background:linear-gradient(135deg,#7c3aed,#6d28d9);color:white;border:none;border-radius:10px;font-size:1rem;font-weight:600;cursor:pointer;margin-top:20px;transition:.2s}}
  .btn:hover{{opacity:.9;transform:translateY(-1px)}}
  .result{{background:#0d1117;border:1px solid #7c3aed33;border-radius:12px;padding:20px;margin-top:16px;display:none}}
  .result-row{{display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #2d2d3f}}
  .result-row:last-child{{border:none}}
  .result-label{{color:#94a3b8;font-size:.9rem}}
  .result-value{{font-weight:700;color:#a78bfa;font-size:1rem}}
  .big-number{{font-size:2.5rem;font-weight:800;text-align:center;background:linear-gradient(135deg,#22c55e,#16a34a);-webkit-background-clip:text;-webkit-text-fill-color:transparent;padding:20px 0}}
  .tips{{background:#1a1a2e;border-left:3px solid #7c3aed;padding:16px;border-radius:0 8px 8px 0;margin-top:12px}}
  .tips h3{{color:#a78bfa;margin-bottom:10px;font-size:.95rem}}
  .tips li{{color:#94a3b8;font-size:.875rem;margin:6px 0;padding-left:16px}}
  .tab-bar{{display:flex;gap:8px;margin-bottom:20px;flex-wrap:wrap}}
  .tab{{padding:8px 16px;background:#2d2d3f;border:none;border-radius:8px;color:#94a3b8;cursor:pointer;font-size:.875rem;transition:.2s}}
  .tab.active{{background:#7c3aed;color:white}}
  .tab-content{{display:none}}.tab-content.active{{display:block}}
</style>
</head>
<body>

<div class="header">
  <h1>⚡ {name}</h1>
  <p class="subtitle">Powered by {topic} intelligence — instant results</p>
</div>

<div class="card" style="max-width:700px;margin:0 auto">
  <div class="tab-bar">
    <button class="tab active" onclick="switchTab('roi')">ROI Calculator</button>
    <button class="tab" onclick="switchTab('goal')">Goal Planner</button>
    <button class="tab" onclick="switchTab('time')">Time Value</button>
    <button class="tab" onclick="switchTab('habit')">Habit Tracker</button>
  </div>

  <!-- ROI Calculator -->
  <div id="tab-roi" class="tab-content active">
    <h2>📈 ROI Calculator</h2>
    <label>Initial Investment ($)</label>
    <input type="number" id="roi-invest" placeholder="e.g. 1000" value="1000">
    <label>Expected Return ($)</label>
    <input type="number" id="roi-return" placeholder="e.g. 1500" value="1500">
    <label>Time Period (months)</label>
    <input type="number" id="roi-months" placeholder="e.g. 12" value="12">
    <button class="btn" onclick="calcROI()">Calculate ROI</button>
    <div class="result" id="roi-result">
      <div class="big-number" id="roi-pct"></div>
      <div class="result-row"><span class="result-label">Net Profit</span><span class="result-value" id="roi-profit"></span></div>
      <div class="result-row"><span class="result-label">Monthly Return</span><span class="result-value" id="roi-monthly"></span></div>
      <div class="result-row"><span class="result-label">Annualized ROI</span><span class="result-value" id="roi-annual"></span></div>
      <div class="tips">
        <h3>💡 ROI Insights</h3>
        <ul><li id="roi-tip1"></li><li id="roi-tip2"></li></ul>
      </div>
    </div>
  </div>

  <!-- Goal Planner -->
  <div id="tab-goal" class="tab-content">
    <h2>🎯 Goal Planner</h2>
    <label>Goal ($)</label>
    <input type="number" id="goal-target" placeholder="e.g. 10000" value="10000">
    <label>Already saved ($)</label>
    <input type="number" id="goal-saved" placeholder="e.g. 2000" value="2000">
    <label>Monthly contribution ($)</label>
    <input type="number" id="goal-monthly" placeholder="e.g. 500" value="500">
    <button class="btn" onclick="calcGoal()">Calculate Timeline</button>
    <div class="result" id="goal-result">
      <div class="big-number" id="goal-months-display"></div>
      <div class="result-row"><span class="result-label">Remaining needed</span><span class="result-value" id="goal-remaining"></span></div>
      <div class="result-row"><span class="result-label">Target date</span><span class="result-value" id="goal-date"></span></div>
      <div class="result-row"><span class="result-label">% Complete</span><span class="result-value" id="goal-pct"></span></div>
    </div>
  </div>

  <!-- Time Value -->
  <div id="tab-time" class="tab-content">
    <h2>⏱️ Hourly Rate Calculator</h2>
    <label>Desired annual income ($)</label>
    <input type="number" id="tv-income" placeholder="e.g. 60000" value="60000">
    <label>Hours worked per week</label>
    <input type="number" id="tv-hours" placeholder="e.g. 40" value="40">
    <label>Vacation weeks per year</label>
    <input type="number" id="tv-vacation" placeholder="e.g. 2" value="2">
    <button class="btn" onclick="calcTime()">Calculate</button>
    <div class="result" id="time-result">
      <div class="big-number" id="time-hourly"></div>
      <div class="result-row"><span class="result-label">Daily rate (8h)</span><span class="result-value" id="time-daily"></span></div>
      <div class="result-row"><span class="result-label">Weekly rate</span><span class="result-value" id="time-weekly"></span></div>
      <div class="result-row"><span class="result-label">Monthly rate</span><span class="result-value" id="time-monthly"></span></div>
    </div>
  </div>

  <!-- Habit Tracker -->
  <div id="tab-habit" class="tab-content">
    <h2>✅ 30-Day Habit Tracker</h2>
    <p style="color:#94a3b8;font-size:.875rem;margin-bottom:16px">Click days to mark as complete. Data saved in browser.</p>
    <label>Habit name</label>
    <input type="text" id="habit-name" placeholder="e.g. Morning workout" value="Morning workout">
    <div id="habit-grid" style="display:grid;grid-template-columns:repeat(7,1fr);gap:6px;margin-top:16px"></div>
    <div class="result" id="habit-result" style="display:block;margin-top:16px">
      <div class="result-row"><span class="result-label">Streak</span><span class="result-value" id="habit-streak">0 days</span></div>
      <div class="result-row"><span class="result-label">Completion rate</span><span class="result-value" id="habit-rate">0%</span></div>
    </div>
  </div>
</div>

<script>
function switchTab(id) {{
  document.querySelectorAll('.tab-content').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('tab-'+id).classList.add('active');
  event.target.classList.add('active');
}}
function fmt(n) {{ return '$'+n.toLocaleString(undefined,{{minimumFractionDigits:2,maximumFractionDigits:2}}); }}
function calcROI() {{
  const inv=+document.getElementById('roi-invest').value;
  const ret=+document.getElementById('roi-return').value;
  const mo=+document.getElementById('roi-months').value;
  if(!inv||!ret||!mo) return;
  const profit=ret-inv, pct=(profit/inv)*100, monthly=profit/mo, annual=((ret/inv)**(12/mo)-1)*100;
  document.getElementById('roi-pct').textContent=pct.toFixed(1)+'% ROI';
  document.getElementById('roi-profit').textContent=fmt(profit);
  document.getElementById('roi-monthly').textContent=fmt(monthly)+'/mo';
  document.getElementById('roi-annual').textContent=annual.toFixed(1)+'%';
  document.getElementById('roi-tip1').textContent=pct>20?'Great return! Above market average (7-10%)':'Consider if this beats index fund returns (~10%/yr)';
  document.getElementById('roi-tip2').textContent='Break-even in '+(inv/monthly).toFixed(1)+' months';
  document.getElementById('roi-result').style.display='block';
}}
function calcGoal() {{
  const target=+document.getElementById('goal-target').value;
  const saved=+document.getElementById('goal-saved').value;
  const monthly=+document.getElementById('goal-monthly').value;
  const remaining=target-saved, months=Math.ceil(remaining/monthly);
  const d=new Date(); d.setMonth(d.getMonth()+months);
  document.getElementById('goal-remaining').textContent=fmt(remaining);
  document.getElementById('goal-months-display').textContent=months+' months';
  document.getElementById('goal-date').textContent=d.toLocaleDateString('en',{{month:'long',year:'numeric'}});
  document.getElementById('goal-pct').textContent=(saved/target*100).toFixed(1)+'%';
  document.getElementById('goal-result').style.display='block';
}}
function calcTime() {{
  const income=+document.getElementById('tv-income').value;
  const hours=+document.getElementById('tv-hours').value;
  const vac=+document.getElementById('tv-vacation').value;
  const weeks=52-vac, hourly=income/(weeks*hours), daily=hourly*8;
  document.getElementById('time-hourly').textContent=fmt(hourly)+'/hr';
  document.getElementById('time-daily').textContent=fmt(daily);
  document.getElementById('time-weekly').textContent=fmt(hourly*hours);
  document.getElementById('time-monthly').textContent=fmt(income/12);
  document.getElementById('time-result').style.display='block';
}}
// Habit tracker
const state=JSON.parse(localStorage.getItem('habit')||'{{}}');
function renderHabit() {{
  const g=document.getElementById('habit-grid'); g.innerHTML='';
  let streak=0,completed=0;
  for(let i=1;i<=30;i++) {{
    const done=state[i];
    if(done) completed++;
    const btn=document.createElement('button');
    btn.textContent=i; btn.onclick=()=>{{state[i]=!state[i];localStorage.setItem('habit',JSON.stringify(state));renderHabit();}};
    btn.style.cssText=`padding:10px;border:1px solid ${{done?'#22c55e':'#2d2d3f'}};background:${{done?'#14532d':'#1e1e2e'}};color:${{done?'#86efac':'#94a3b8'}};border-radius:8px;cursor:pointer;font-size:.875rem;`;
    g.appendChild(btn);
  }}
  for(let i=30;i>=1;i--) {{ if(state[i]) streak++; else break; }}
  document.getElementById('habit-streak').textContent=streak+' days';
  document.getElementById('habit-rate').textContent=(completed/30*100).toFixed(0)+'%';
}}
renderHabit();
</script>
</body>
</html>"""

    html_path = product_dir / f"{pid}_tools.html"
    html_path.write_text(html, encoding="utf-8")

    readme = f"""# {name} — Interactive Tools

## Open in any browser. No installation required.

1. Download `{pid}_tools.html`
2. Double-click to open in Chrome/Safari/Firefox
3. All data is saved locally in your browser

## 4 Tools Included

| Tool | What it does |
|---|---|
| ROI Calculator | Calculate return on any investment |
| Goal Planner | Know exactly when you'll hit your savings goal |
| Hourly Rate | Find your minimum billable rate |
| Habit Tracker | 30-day visual tracker, persists in browser |

*Works offline. No account required. © AETERNA Holdings*
"""
    (product_dir / "README.md").write_text(readme, encoding="utf-8")

    zip_path = product_dir / f"{pid}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(html_path, html_path.name)
        zf.write(product_dir / "README.md", "README.md")
    return zip_path


# ─── Checklist Pack ──────────────────────────────────────────────────────────

def create_checklist_pack(spec: dict, product_dir: Path):
    name  = spec["name"]
    pid   = spec["product_id"]
    topic = spec.get("subcategory", "Business")

    launch_checklist = f"""# {name}
## The Ultimate {topic} Checklist Pack

> 5 comprehensive checklists. 200+ action items. Print-ready or use digitally.

---

## CHECKLIST 1: Business Launch Checklist (50 items)

### Legal & Structure
- [ ] Choose business structure (LLC / Sole Prop / Corp)
- [ ] Register business name
- [ ] Get EIN (Employer Identification Number) from IRS
- [ ] Open separate business bank account
- [ ] Get business insurance
- [ ] Create operating agreement / bylaws
- [ ] Register trademarks if needed
- [ ] Set up accounting software (QuickBooks / Wave)

### Financial Setup
- [ ] Set startup budget
- [ ] Define pricing model (hourly / project / retainer / subscription)
- [ ] Create invoice template
- [ ] Set up payment processing (Stripe / PayPal / Square)
- [ ] Open business credit card
- [ ] Set aside 25-30% for taxes from day 1
- [ ] Create 6-month runway emergency fund
- [ ] Set up bookkeeping system (weekly habit)

### Brand & Online Presence
- [ ] Define core brand values (3-5)
- [ ] Choose brand colors + fonts
- [ ] Get logo (Canva / Fiverr / designer)
- [ ] Buy domain name
- [ ] Set up professional email (yourname@yourdomain.com)
- [ ] Build website / landing page
- [ ] Set up Google Business Profile
- [ ] Create LinkedIn company page
- [ ] Set up relevant social accounts
- [ ] Create email list (Mailchimp / ConvertKit free tier)

### Product / Service
- [ ] Define your core offer clearly (what + who + result)
- [ ] Write offer description (100 words or fewer)
- [ ] Set pricing (research competitors first)
- [ ] Create sales page or pitch deck
- [ ] Build onboarding process for new clients
- [ ] Create service agreement / contract template
- [ ] Define refund / cancellation policy
- [ ] Build FAQ document
- [ ] Get 3 beta testers / early clients

### Marketing & Sales
- [ ] Identify top 3 customer acquisition channels
- [ ] Create lead magnet (free resource)
- [ ] Write 5 email templates (intro / follow-up / proposal / onboarding / offboarding)
- [ ] Set up CRM (HubSpot free / Notion)
- [ ] Define referral program
- [ ] Plan first 30 days of content
- [ ] Reach out to 50 warm contacts
- [ ] List on relevant directories / marketplaces
- [ ] Join 3 niche communities / forums
- [ ] Partner with 2 complementary businesses

### Operations
- [ ] Create standard operating procedures (SOPs) for core processes
- [ ] Set up project management tool (Notion / Trello / Asana)
- [ ] Define communication channels (Slack / email rules)
- [ ] Set working hours and availability
- [ ] Set up automated backups for files
- [ ] Create password manager for business accounts
- [ ] Set recurring calendar events for weekly/monthly reviews

---

## CHECKLIST 2: Website Launch Checklist (40 items)

### Before Launch
- [ ] Define website goal (leads / sales / portfolio / info)
- [ ] Map site structure (pages list)
- [ ] Write all copy before designing
- [ ] Collect all images / assets
- [ ] Set up Google Analytics 4
- [ ] Set up Google Search Console
- [ ] Install SSL certificate (HTTPS)
- [ ] Test on mobile (iPhone + Android)
- [ ] Test on multiple browsers (Chrome / Safari / Firefox)
- [ ] Check page load speed (target <3 seconds) via PageSpeed Insights

### SEO Checklist
- [ ] Research 10 target keywords
- [ ] Write meta title for each page (50-60 chars)
- [ ] Write meta description for each page (150-160 chars)
- [ ] Add alt text to all images
- [ ] Create XML sitemap
- [ ] Submit sitemap to Google Search Console
- [ ] Set up canonical URLs
- [ ] Check for broken links
- [ ] Add schema markup to key pages
- [ ] Get 5 backlinks before launch

### Content & Copy
- [ ] Homepage: clear H1 headline (value proposition)
- [ ] Homepage: social proof (testimonials / logos / numbers)
- [ ] Homepage: clear CTA above the fold
- [ ] About page: founder story + credibility
- [ ] Contact page: multiple contact options
- [ ] Privacy policy page
- [ ] Terms of service page
- [ ] Cookie consent banner (EU/GDPR)
- [ ] 404 page with navigation
- [ ] Blog / resources section set up

### Technical
- [ ] Forms tested and working
- [ ] Email notifications configured
- [ ] Payment system tested (if applicable)
- [ ] CDN configured for fast global delivery
- [ ] Image compression done
- [ ] Minified CSS/JS
- [ ] Robots.txt file in place
- [ ] Backup system in place
- [ ] Hosting plan matches expected traffic
- [ ] Uptime monitoring set up (UptimeRobot free)

---

## CHECKLIST 3: Content Marketing Checklist (30 items)

### Strategy
- [ ] Define content goal (awareness / leads / sales)
- [ ] Choose 2-3 primary channels (blog / YouTube / social / email)
- [ ] Research top 20 content topics in your niche
- [ ] Create content calendar (12 weeks minimum)
- [ ] Define content pillars (3-5 core themes)
- [ ] Batch-create content (at least 4 pieces ahead)

### Production
- [ ] Create content templates for each format
- [ ] Set up Canva templates for graphics
- [ ] Set up recording setup (webcam / mic) if video
- [ ] Create brand hashtag list (primary / secondary / niche)
- [ ] Write 10 headlines before picking the best one
- [ ] Add strong hook in first 3 seconds (video) or first line (text)
- [ ] Include clear CTA in every piece of content
- [ ] Repurpose each piece into 3+ formats

### Distribution
- [ ] Post at optimal times (test, then standardize)
- [ ] Cross-post to all relevant platforms
- [ ] Email newsletter every week
- [ ] Pin best content on profiles
- [ ] Engage first 30 min after posting (respond to comments)
- [ ] DM new followers who engage

### Analytics
- [ ] Track weekly: followers / reach / engagement rate
- [ ] Track monthly: leads / sales from content
- [ ] Identify top 20% performing content → double down
- [ ] A/B test headlines / thumbnails
- [ ] Review and update old top-performing content

---

## CHECKLIST 4: Freelance Client Onboarding (30 items)

### Before the Project
- [ ] Discovery call complete
- [ ] Project scope documented in writing
- [ ] Quote/proposal sent and approved
- [ ] Contract signed (both parties)
- [ ] 50% deposit received before work starts
- [ ] Project timeline agreed
- [ ] Communication preferences established (email / Slack / calls)
- [ ] Access to tools/accounts granted
- [ ] Meeting cadence scheduled

### During the Project
- [ ] Weekly status update sent every Friday
- [ ] Deliverables shared via agreed channel (not email attachment)
- [ ] Feedback rounds defined (max 2 rounds)
- [ ] Change requests documented and priced
- [ ] Questions batched (not one at a time)
- [ ] Progress tracked in shared doc

### Project Wrap-Up
- [ ] All deliverables sent and confirmed received
- [ ] Final invoice sent
- [ ] Payment received
- [ ] Testimonial/review requested
- [ ] Referral asked for
- [ ] Offboarding document sent (how to maintain / next steps)
- [ ] Archive all project files
- [ ] Add to portfolio (with permission)
- [ ] Send 30-day check-in email

---

## CHECKLIST 5: Daily Productivity Checklist

### Morning (Do Before Email)
- [ ] Identify your #1 task for today (MIT — Most Important Task)
- [ ] Block 90 min deep work time on calendar
- [ ] Review yesterday's open loops
- [ ] Quick physical movement (5-10 min)
- [ ] No social media first 30 minutes

### Deep Work Session
- [ ] Close all tabs except the one you need
- [ ] Phone on Do Not Disturb
- [ ] Timer set (Pomodoro: 25 min on / 5 min break)
- [ ] Single-task only
- [ ] Note distracting thoughts → parking lot list

### End of Day
- [ ] Complete open tasks or re-schedule them
- [ ] Clear inbox to zero (or sort to action folders)
- [ ] Prep tomorrow's MIT
- [ ] Note one win from today
- [ ] Shut down ritual (close laptop, say "shutdown complete")

---
*© AETERNA Holdings — Print, copy, share freely.*
"""

    (product_dir / "checklists.md").write_text(launch_checklist, encoding="utf-8")

    html_checklist = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} — Interactive Checklists</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,sans-serif;background:#f8fafc;color:#1e293b;padding:20px}}
.container{{max-width:800px;margin:0 auto}}
h1{{font-size:1.8rem;font-weight:700;color:#1e293b;margin-bottom:4px}}
.subtitle{{color:#64748b;margin-bottom:24px;font-size:.95rem}}
.section{{background:white;border-radius:12px;padding:20px;margin-bottom:16px;border:1px solid #e2e8f0}}
.section h2{{font-size:1.1rem;font-weight:600;color:#7c3aed;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #f1f5f9}}
.item{{display:flex;align-items:flex-start;gap:10px;padding:6px 0;cursor:pointer}}
.item input[type=checkbox]{{width:16px;height:16px;margin-top:2px;accent-color:#7c3aed;cursor:pointer;flex-shrink:0}}
.item label{{cursor:pointer;font-size:.9rem;color:#374151;transition:.2s;line-height:1.4}}
.item.done label{{text-decoration:line-through;color:#94a3b8}}
.progress{{background:#f1f5f9;border-radius:4px;height:6px;margin-bottom:16px}}
.progress-bar{{background:linear-gradient(90deg,#7c3aed,#ec4899);height:100%;border-radius:4px;transition:.3s}}
.stats{{display:flex;gap:16px;font-size:.8rem;color:#64748b;margin-bottom:12px}}
.btn-reset{{background:#fee2e2;color:#dc2626;border:none;padding:6px 12px;border-radius:6px;cursor:pointer;font-size:.8rem;margin-left:auto}}
</style>
</head>
<body>
<div class="container">
<h1>✅ {name}</h1>
<p class="subtitle">Interactive checklist — progress saved in your browser</p>
<div id="app"></div>
</div>
<script>
const lists=[
  {{title:"🚀 Business Launch (50 items)",key:"launch",items:["Choose business structure","Register business name","Get EIN / Tax ID","Open business bank account","Get business insurance","Create operating agreement","Open business credit card","Set up accounting software","Define pricing model","Create invoice template","Set up payment processing","Set startup budget","Buy domain name","Set up professional email","Build website","Set up Google Business Profile","Create LinkedIn page","Set up social accounts","Create email list","Design brand colors & logo","Define core offer","Write offer description (100 words)","Set pricing after research","Create sales page","Build onboarding process","Create service contract template","Define refund policy","Build FAQ document","Get 3 beta clients","Identify top 3 acquisition channels","Create lead magnet","Write 5 email templates","Set up CRM","Define referral program","Plan 30-day content calendar","Reach out to 50 warm contacts","List on relevant directories","Join 3 niche communities","Partner with complementary businesses","Create SOPs for core processes","Set up project management tool","Define communication channels","Set working hours","Set up automated backups","Create password manager","Set weekly review habit","Track key metrics weekly","Apply for business grants","Get business mentor","Celebrate your launch!"]}},
  {{title:"🌐 Website Launch (20 items)",key:"website",items:["Define website goal","Map site structure","Write all copy first","Set up Google Analytics 4","Install SSL certificate","Test on mobile","Test page load speed","Research 10 target keywords","Write meta titles for all pages","Write meta descriptions","Add alt text to images","Create XML sitemap","Test all forms","Homepage: clear headline","Homepage: social proof","Homepage: CTA above fold","Privacy policy page","Terms of service page","404 page set up","Backup system in place"]}},
  {{title:"📱 Content Marketing (20 items)",key:"content",items:["Define content goal","Choose 2-3 primary channels","Research top 20 content topics","Create 12-week content calendar","Define content pillars","Batch create 4+ pieces ahead","Create design templates","Set up brand hashtag list","Write 10 headlines, pick best","Add hook in first line","Include CTA in every piece","Repurpose content to 3 formats","Post at optimal times","Email newsletter weekly","Track reach & engagement weekly","Identify top performing content","A/B test headlines","Update old top content","Engage 30 min after posting","DM new followers who engage"]}},
  {{title:"💼 Daily Productivity (15 items)",key:"daily",items:["Identify #1 task (MIT)","Block 90-min deep work","Close distracting tabs","Phone on Do Not Disturb","No social media first 30 min","Use Pomodoro timer (25/5)","Single-task only","Note distractions, don't follow","Complete or reschedule open tasks","Clear inbox to zero","Prep tomorrow's MIT","Note one win from today","Shut down ritual (close laptop)","Drink 8 glasses water","Get 7-8 hours sleep"]}}
];

function save(){{localStorage.setItem('cl',JSON.stringify(state));}}
const state=JSON.parse(localStorage.getItem('cl')||'{{}}');

function renderAll(){{
  const app=document.getElementById('app'); app.innerHTML='';
  lists.forEach(list=>{{
    const done=(state[list.key]||[]).length;
    const pct=Math.round(done/list.items.length*100);
    const div=document.createElement('div'); div.className='section';
    div.innerHTML=`<h2>${{list.title}}</h2>
    <div class="stats"><span>${{done}}/${{list.items.length}} done</span><span>${{pct}}% complete</span><button class="btn-reset" onclick="resetList('${{list.key}}')">Reset</button></div>
    <div class="progress"><div class="progress-bar" style="width:${{pct}}%"></div></div>`;
    list.items.forEach((item,i)=>{{
      const checked=(state[list.key]||[]).includes(i);
      const row=document.createElement('div'); row.className='item'+(checked?' done':'');
      row.innerHTML=`<input type="checkbox" id="${{list.key}}-${{i}}" ${{checked?'checked':''}}><label for="${{list.key}}-${{i}}">${{item}}</label>`;
      row.querySelector('input').onchange=e=>{{
        if(!state[list.key]) state[list.key]=[];
        if(e.target.checked) state[list.key].push(i);
        else state[list.key]=state[list.key].filter(x=>x!==i);
        save(); renderAll();
      }};
      div.appendChild(row);
    }});
    app.appendChild(div);
  }});
}}
function resetList(key){{if(confirm('Reset this checklist?')){{state[key]=[];save();renderAll();}}}}
renderAll();
</script>
</body>
</html>"""

    html_path = product_dir / f"{pid}_interactive.html"
    html_path.write_text(html_checklist, encoding="utf-8")

    zip_path = product_dir / f"{pid}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(product_dir / "checklists.md", "checklists.md")
        zf.write(html_path, html_path.name)
    return zip_path


# ─── Business Template Pack ───────────────────────────────────────────────────

def create_business_template(spec: dict, product_dir: Path):
    name  = spec["name"]
    pid   = spec["product_id"]

    freelance_contract = f"""# FREELANCE SERVICE AGREEMENT

**This Agreement** is made between:

**Client:** _____________________________ ("Client")
**Freelancer:** _________________________ ("Freelancer")
**Date:** _______________________________

---

## 1. SERVICES

Freelancer agrees to provide the following services:

**Project:** ___________________________
**Description:**
_________________________________________
_________________________________________

**Deliverables:**
- [ ] ____________________________
- [ ] ____________________________
- [ ] ____________________________

---

## 2. TIMELINE

| Milestone | Deliverable | Due Date |
|---|---|---|
| Kickoff | Project brief review | Day 1 |
| Draft | First draft / prototype | Day ___ |
| Revisions | Client feedback round 1 | Day ___ |
| Final | Final delivery | Day ___ |

**Start Date:** _______________
**End Date:** _________________

---

## 3. COMPENSATION

**Total Project Fee:** $___________

**Payment Schedule:**
- 50% deposit due before work begins: $__________
- 50% balance due upon completion: $__________

**Payment Method:** _________________ (PayPal / Bank transfer / Stripe)
**Payment Terms:** Net 7 days from invoice

**Late Payment:** Invoices unpaid after 14 days incur 1.5% monthly interest.

---

## 4. REVISIONS

This agreement includes **___ rounds** of revisions.
Additional revisions are billed at $___/hour.

**Revision requests must be submitted within 7 days** of receiving deliverables.

---

## 5. INTELLECTUAL PROPERTY

Upon receipt of full payment, Client receives full ownership of the final deliverables.

Freelancer retains the right to display the work in their portfolio.

Freelancer retains ownership of any pre-existing materials, tools, or methodologies used.

---

## 6. CONFIDENTIALITY

Freelancer agrees to keep all Client information confidential and not disclose it to third parties without written consent.

---

## 7. CANCELLATION

- **Client cancels before work starts:** Full deposit refunded
- **Client cancels after work starts:** Deposit is forfeited; Client pays for work completed
- **Freelancer cancels:** Full deposit refunded; Freelancer completes deliverables already started

---

## 8. LIMITATION OF LIABILITY

Freelancer's total liability is limited to the amount paid for the project.
Freelancer is not liable for indirect, consequential, or special damages.

---

## 9. DISPUTE RESOLUTION

Both parties agree to attempt resolution through direct communication first.
If unresolved, disputes will be settled by binding arbitration in [State/Country].

---

## 10. ENTIRE AGREEMENT

This document constitutes the entire agreement between the parties. Any changes must be agreed in writing.

---

**CLIENT SIGNATURE:** _________________________ Date: ________

**FREELANCER SIGNATURE:** ___________________ Date: ________

---
*Template provided by AETERNA Holdings. Consult a lawyer for your jurisdiction.*
"""

    invoice_template = f"""# INVOICE

**INVOICE #:** INV-[YYYY]-[XXXX]
**DATE:** _______________
**DUE DATE:** _______________ (Net ___ days)

---

**FROM (Freelancer):**
Name: ___________________________
Email: __________________________
Phone: __________________________
Address: ________________________
Tax ID / EIN: ___________________

---

**TO (Client):**
Company: ________________________
Contact: ________________________
Email: __________________________
Address: ________________________

---

## Services Rendered

| Description | Qty | Rate | Amount |
|---|---|---|---|
| ________________________ | ___ | $_____ | $_____ |
| ________________________ | ___ | $_____ | $_____ |
| ________________________ | ___ | $_____ | $_____ |

---

**Subtotal:** $__________
**Tax (___ %):** $__________
**Discount:** -$__________
**TOTAL DUE:** $__________

---

**Payment Methods:**
- PayPal: ____________________
- Bank Transfer: ____________________  (include invoice # in reference)
- Venmo / Zelle: ____________________

**Payment Terms:** Due within ___ days of invoice date.
**Late Fees:** 1.5% per month on overdue balances.

---

**Notes:**
_________________________________________

Thank you for your business!

---
*Template by AETERNA Holdings*
"""

    sop_template = f"""# STANDARD OPERATING PROCEDURE (SOP)

**Process Name:** ___________________________
**Department / Owner:** _____________________
**Version:** 1.0
**Last Updated:** ___________________________
**Approved By:** ____________________________

---

## 1. PURPOSE

What this SOP achieves:
_________________________________________

When to use this SOP:
_________________________________________

---

## 2. SCOPE

**Applies to:** ____________________________
**Does not apply to:** _____________________

---

## 3. RESPONSIBILITIES

| Role | Responsibility |
|---|---|
| ______________ | ______________________________ |
| ______________ | ______________________________ |
| ______________ | ______________________________ |

---

## 4. TOOLS REQUIRED

- [ ] ___________________________
- [ ] ___________________________
- [ ] ___________________________

---

## 5. PROCEDURE (Step by Step)

### Step 1: [Name]
**Who:** _______________
**When:** ______________
**How:**
1. ____________________
2. ____________________
3. ____________________
**Output:** ____________

### Step 2: [Name]
**Who:** _______________
**When:** ______________
**How:**
1. ____________________
2. ____________________
3. ____________________
**Output:** ____________

### Step 3: [Name]
(Continue as needed...)

---

## 6. QUALITY CHECKS

Before marking complete, verify:
- [ ] ___________________________
- [ ] ___________________________
- [ ] ___________________________

---

## 7. TROUBLESHOOTING

| Problem | Solution |
|---|---|
| ______________ | ______________________________ |
| ______________ | ______________________________ |

---

## 8. REVISION HISTORY

| Version | Date | Changes | Author |
|---|---|---|---|
| 1.0 | _______ | Initial version | _______ |

---
*Template by AETERNA Holdings*
"""

    project_proposal = f"""# PROJECT PROPOSAL

**Prepared by:** ___________________________
**Prepared for:** __________________________
**Date:** ___________________________________
**Proposal #:** ____________________________
**Valid Until:** ____________________________

---

## EXECUTIVE SUMMARY

We propose to help [Client] achieve [Goal] by delivering [Solution].
This project will take [Timeline] and requires an investment of $[Amount].

---

## THE PROBLEM

**Current situation:**
_________________________________________

**Pain points:**
- ___________________________
- ___________________________
- ___________________________

**Cost of inaction (estimate):** $__________ per month

---

## OUR SOLUTION

**What we'll deliver:**
_________________________________________

**Our approach:**
1. Phase 1: _______________________
2. Phase 2: _______________________
3. Phase 3: _______________________

---

## SCOPE OF WORK

### Included:
- [ ] ___________________________
- [ ] ___________________________
- [ ] ___________________________

### Not included:
- [ ] ___________________________
- [ ] ___________________________

---

## TIMELINE

| Week | Milestone | Deliverable |
|---|---|---|
| 1 | Kickoff | Discovery complete |
| 2-3 | Phase 1 | ________________ |
| 4-5 | Phase 2 | ________________ |
| 6 | Delivery | Final deliverables |

**Start date:** _______________
**End date:** _________________

---

## INVESTMENT

| Item | Amount |
|---|---|
| _________________ | $_______ |
| _________________ | $_______ |
| **TOTAL** | **$_______** |

**Payment schedule:**
- 50% deposit: $_______ (due to start)
- 50% on delivery: $_______ (due on completion)

---

## WHY US

- Years of experience: ___
- Relevant clients: ___________________________
- Our superpower: ___________________________

**Results we've achieved:**
- ___________________________
- ___________________________

---

## NEXT STEPS

1. Review this proposal
2. Reply with questions / requested changes
3. Sign agreement + pay deposit
4. Kick-off call scheduled

**Questions?** Contact us at: ___________________________

---

*This proposal is valid for 14 days from the date above.*
*© AETERNA Holdings Template*
"""

    (product_dir / "freelance-contract.md").write_text(freelance_contract, encoding="utf-8")
    (product_dir / "invoice-template.md").write_text(invoice_template, encoding="utf-8")
    (product_dir / "sop-template.md").write_text(sop_template, encoding="utf-8")
    (product_dir / "project-proposal.md").write_text(project_proposal, encoding="utf-8")

    index = f"""# {name} — Business Template Pack

## What's Included

| File | Use Case |
|---|---|
| freelance-contract.md | Full freelance service agreement |
| invoice-template.md | Professional invoice template |
| sop-template.md | Standard operating procedure |
| project-proposal.md | Client proposal template |

## How to Use

1. Open any template in a text editor or Notion
2. Fill in the [BRACKETS] with your information
3. For contracts: copy into Google Docs for easy editing and signing
4. For invoices: copy into Google Sheets or use as-is

## Formats

All templates are in Markdown — universal format that works in:
- **Notion** (paste directly)
- **Google Docs** (paste → Format → Convert to Docs)
- **Obsidian**
- Any text editor

*© AETERNA Holdings — Use for personal and commercial projects.*
"""
    (product_dir / "README.md").write_text(index, encoding="utf-8")

    zip_path = product_dir / f"{pid}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in product_dir.glob("*.md"):
            zf.write(f, f.name)
    return zip_path


# ─── Dispatcher ──────────────────────────────────────────────────────────────

TYPE_MAP = {
    "notion_template":    create_notion_template,
    "prompt_pack":        create_prompt_pack,
    "script_pack":        create_script_pack,
    "guide":              create_guide,
    "ebook":              create_guide,
    "toolkit":            create_script_pack,
    "template":           create_notion_template,
    "course":             create_guide,
    "spreadsheet_pack":   create_spreadsheet_pack,
    "html_tool":          create_html_tool,
    "checklist_pack":     create_checklist_pack,
    "business_template":  create_business_template,
}

def create_product(spec: dict, product_dir: Path) -> dict:
    ptype   = spec.get("product_type", "guide")
    creator = TYPE_MAP.get(ptype, create_guide)
    creator(spec, product_dir)

    # content.md が空/短ければ guide で補完
    cm = product_dir / "content.md"
    if not cm.exists() or cm.stat().st_size < 500:
        create_guide(spec, product_dir)

    # gumroad_page.md を標準フォーマットで生成
    price = spec.get("price_usd", 19)
    topic = spec.get("subcategory", "")
    (product_dir / "gumroad_page.md").write_text(
        f"""## Description

### {spec['name']}

**What you get:** A complete, actionable {ptype} on {topic}.

Focused. No fluff. Immediate value.

### What's Inside
- Step-by-step practical content
- Real examples you can use today
- Templates and frameworks included
- Beginner to intermediate friendly

### One-time purchase. Instant download.

**${price}** — 30-day money-back guarantee.
""", encoding="utf-8")

    return {"product_id": spec["product_id"], "name": spec["name"], "type": ptype,
            "files": [f.name for f in product_dir.iterdir() if f.is_file()]}
